import glob

import openpyxl
import pandas as pd

from scraping import excel_tabling


def summary(month, file_path, files):
    print(file_path)

    # ベース用データフレームの準備
    df_base = pd.read_excel(files[0], index_col=None, header=0)

    for file in files[1:]:
        print(file)

        # 追加用データフレームの準備
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        data = []
        for i in range(2, ws.max_row):
            datum = {
                '検索結果数': ws[f'A{i}'].value,
                '抽出日時': ws[f'B{i}'].value,
                'thread_ts': ws[f'C{i}'].value,
                '投稿日': ws[f'D{i}'].value,
                '投稿時間': ws[f'E{i}'].value,
                '投稿チャンネル': ws[f'F{i}'].value,
                '投稿者': ws[f'G{i}'].value,
                '投稿メッセージ': ws[f'H{i}'].value,
                'リンク': ws[f'I{i}'].value,
                'リンク2': ws[f'H{i}'].value,
                'リンク3': ws[f'J{i}'].value,
                'リンク4': ws[f'K{i}'].value,
                'リンク5': ws[f'L{i}'].value,
                'リンク6': ws[f'M{i}'].value
            }
            data.append(datum)

        _df = pd.DataFrame(data)
        _df = _df.dropna(how='all')

        # 2つのデータフレームを縦に結合し新たなベース用データフレームとする
        df_base = pd.concat([df_base, _df])

    return df_base, file_path


if __name__ == '__main__':

    # 2022年3月～2022年7月の集計
    for month in range(3, 8):
        file_path = f'src/2022-{month:02}'
        files = glob.glob(f'{file_path}/*_c4b_slack.xlsx')

        print(f'{month}月の集計を開始')
        df_base, file_path = summary(month, file_path, files)

        df_base.to_excel(f'{file_path}/月集計.xlsx', index=False)
        excel_tabling(f'{file_path}/月集計.xlsx')

    # 月ごとを更に1本化
    file_path = f'src'
    files = glob.glob(f'{file_path}/*/月集計.xlsx')

    print(file_path)
    df_base, file_path = summary(month, file_path, files)

    df_base.to_excel(f'{file_path}/集計.xlsx', index=False)
    excel_tabling(f'{file_path}/集計.xlsx')
    print(f'{file_path}/集計.xlsx へ出力完了しました')
