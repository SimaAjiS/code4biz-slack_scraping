import json
import math
from time import sleep

import openpyxl
import pandas as pd
import pyautogui as pg
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


def selenium_run():
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    url = 'https://code4biz.slack.com/'

    # 暗黙的な待機
    driver.implicitly_wait(5)
    driver.get(url)
    driver.maximize_window()
    return driver


def auto_login(driver):
    # ログインページ処理
    form = driver.find_element(by=By.CSS_SELECTOR, value='#signin_form')
    login_mail = form.find_element(by=By.NAME, value='email')
    login_passwd = form.find_element(by=By.NAME, value='password')

    # 認証情報
    with open('src/auth.json') as f:
        auth = json.load(f)

    my_mail = auth['my_mail']
    my_pass = auth['my_pass']

    # ログイン情報初期化
    login_mail.clear()
    login_passwd.clear()

    login_mail.send_keys(my_mail)
    login_passwd.send_keys(my_pass)

    # サインインボタンを押す
    sleep(0.5)
    btn = form.find_element(by=By.CSS_SELECTOR, value='#signin_btn')
    btn.click()
    sleep(2)

    # Slackアプリ起動確認メッセージのキャンセル
    pg.press('enter', presses=1, interval=0.1)
    sleep(1)

    # ブラウザでSlack起動
    pg.press('tab', presses=2, interval=0.1)
    pg.press('enter', presses=1, interval=0.1)

    # SlackのWebページが完全に立ち上がるまで8秒待機
    sleep(8)


def excel_tabling(file_name):
    # ワークブックを開く
    wb = openpyxl.load_workbook(file_name, data_only=True)

    # 最後のシートを選択
    ws_summary = wb.worksheets[-1]

    # テーブルを生成する
    max_row = ws_summary.max_row
    table = Table(displayName='Tメンバー', ref=f'A1:B{max_row}')

    # テーブルのスタイルを決める(紫、テーブルスタイル(中間))
    table_style = TableStyleInfo(name='TableStyleMedium5', showRowStripes=True)

    # テーブルのスタイルを設定
    table.tableStyleInfo = table_style

    # シートにテーブルを追加
    ws_summary.add_table(table)

    # テーブル化した列幅の調整
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 30

    # テーブル化したフォントの設定
    font_name = 'Meiryo'

    # タイトル行のみ文字色を白色へ
    for i in range(1, ws_summary.max_row + 1):
        ws_summary.cell(row=1, column=i).font = Font(name=font_name, sz=11, color='FFFFFF')

    for i in range(2, ws_summary.max_row + 1):
        ws_summary.cell(row=i, column=1).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=2).font = Font(name=font_name, sz=11)

    wb.save(file_name)


def main(save_file_name):
    driver = selenium_run()
    auto_login(driver)

    # メンバーディレクトリを表示
    pg.hotkey('ctrl', 'shift', 'e')
    sleep(5)

    # 全員表示
    pg.press('tab', presses=1, interval=0.1)
    pg.press('enter', presses=1, interval=0.1)
    pg.press('down', presses=2, interval=0.1)
    pg.press('enter', presses=1, interval=0.1)
    sleep(2)

    # 50件/ページ として総ページ算出
    pages = math.ceil(110 / 50)
    print(f'計 {pages}ページ')

    data = []
    for page in range(pages):
        # メンバーのXPATH取得
        members_grid = driver.find_elements(by=By.CLASS_NAME, value='p-bp__grid_cell')
        print(len(members_grid))
        sleep(1)

        # 1人のメンバー取り出す
        for i, member_grid in enumerate(members_grid):
            member_grid.click()
            sleep(1)

            # 表示名の取得
            profile_name = driver.find_element(by=By.CLASS_NAME, value='p-r_member_profile__name__text').text

            # 連絡先取得
            email = driver.find_element(by=By.CLASS_NAME, value='c-link').get_attribute('href').split(':')[-1]

            # @XXXX名の取得
            buttons = driver.find_elements(by=By.CLASS_NAME, value='p-member_profile_buttons__button_body')
            buttons[-1].click()
            sleep(1)
            mention_name = driver.find_elements(by=By.CLASS_NAME, value='c-menu_item__label')[0].text.split(': ')[-1]
            driver.find_elements(by=By.CLASS_NAME, value='c-menu_item__label')[0].click()

            datum = {
                'profile_name': profile_name,
                'mention_name': mention_name
                # 'email': email
            }

            data.append(datum)
            print(f'{page + 1}/{pages}ページ {i + 1} {profile_name}さんのデータ取得完了')
            sleep(0.5)

        print(f'{page + 1}/{pages}ページの取得完了しました')

        if page + 1 < pages:
            # 一番下まで移動
            pg.press('end', presses=3, interval=1)  # 大抵は1回で下へ遷移するが、たまに動かないため3回
            sleep(2)
            try:
                next_button = driver.find_element(by=By.XPATH, value='//*[@aria-label="次のページ"]')
                next_button.click()
                print(f'\n{page + 2}へのボタンをクリックしました')
            except Exception as e:
                print(e)
        else:
            pass
        sleep(2)

    df = pd.DataFrame(data)
    df.to_excel(save_file_name, index=False)
    excel_tabling(save_file_name)
    driver.quit()


if __name__ == '__main__':
    # 保存ファイル名の設定
    save_file_name = 'src/member_profile.xlsx'

    main(save_file_name)