import json
import math
import os
from datetime import datetime as dt
from datetime import timedelta
from time import sleep

import openpyxl
import pandas as pd
import pyautogui as pg
import pyperclip
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager


def create_search_day_list(start, end):
    # 日付条件の設定
    strdt = dt.strptime(start, '%Y-%m-%d')
    enddt = dt.strptime(end, '%Y-%m-%d')

    # 日付差の日数を算出（リストに最終日も含めたいので、＋１しています）
    days_num = (enddt - strdt).days + 1

    search_days = []
    for i in range(days_num):
        search_days.append(strdt + timedelta(days=i))

    return search_days


def create_src_dir(search_day):  # '2022-03-21'
    year = int(search_day.split('-')[0])
    month = int(search_day.split('-')[1])

    src_dir = f'src/{year}-{month:02}'
    os.makedirs(src_dir, exist_ok=True)
    return src_dir


def selenium_run():
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service)
    url = 'https://code4biz.slack.com/'

    # 暗黙的な待機
    driver.implicitly_wait(5)
    driver.get(url)
    driver.maximize_window()
    return driver


def auto_login(driver):
    # ログインページ処理
    form = driver.find_element(by=By.CSS_SELECTOR, value='#signin_form')
    login_mail = form.find_element(by=By.NAME, value='email')
    login_passwd = form.find_element(by=By.NAME, value='password')

    # 認証情報
    with open('src/auth.json') as f:
        auth = json.load(f)

    my_mail = auth['my_mail']
    my_pass = auth['my_pass']

    # ログイン情報初期化
    login_mail.clear()
    login_passwd.clear()

    login_mail.send_keys(my_mail)
    login_passwd.send_keys(my_pass)

    # サインインボタンを押す
    sleep(0.5)
    btn = form.find_element(by=By.CSS_SELECTOR, value='#signin_btn')
    btn.click()
    sleep(2)

    # Slackアプリ起動確認メッセージのキャンセル
    pg.press('enter', presses=1, interval=0.1)
    sleep(1)

    # ブラウザでSlack起動
    pg.press('tab', presses=2, interval=0.1)
    pg.press('enter', presses=1, interval=0.1)

    # SlackのWebページが完全に立ち上がるまで5秒待機
    sleep(8)


def jump_to_search_box():
    # 検索フィルターで期間指定
    pg.press('tab', presses=12, interval=0.2)
    pg.press('enter', presses=1, interval=0.1)
    sleep(1)


def to_search_date(search_day):
    jump_to_search_box()

    # 日時指定
    pg.press('tab', presses=4, interval=0.2)
    pg.press('enter', presses=1, interval=0.1)
    pg.press('up', presses=4, interval=0.2)
    pg.press('enter', presses=1, interval=0.1)

    # 開始日時を”YYYY-MM-DD”形式で入力
    write_jp(search_day)

    # 保存ボタンを押す
    sleep(0.5)
    pg.hotkey('shift', 'tab')
    pg.hotkey('shift', 'tab')
    pg.press('enter', presses=1, interval=0.1)

    # 検索開始（上記保存することでコメント欄にカーソルアクティブしている）
    sleep(0.2)
    pg.press('tab', presses=15, interval=0.2)
    pg.press('enter', presses=1, interval=0.1)

    print(f'検索日:{search_day}')


# PyAutoGUIで日本語入力も対応する
def write_jp(text):
    sleep(0.5)
    pyperclip.copy(text)
    pg.hotkey('ctrl', 'v')


def display_smaller():
    # 検索結果表示画面の表示を小さくする（一度の表示できる件数を増やす）
    sleep(1)
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    pg.hotkey('ctrl', '-')
    sleep(2)


def get_text_section(message, class_name):
    text_section = message.find_element(by=By.CLASS_NAME, value=class_name).text
    try:
        # リスト表記をしているメッセージへの対応
        lists = message.find_elements(by=By.TAG_NAME, value='li')
        if lists is not None:
            for list in lists:
                text_section += list.text + '\n'
    except:
        pass
    return text_section


# course_link列をハイパーリンク化
def make_clickable(course_link):
    return f'<a target="_blank" href="{course_link}">{course_link}</a>'


def excel_tabling(src_dir, search_day):
    # ワークブックを開く
    aggre_file_name = f'{src_dir}/{search_day}_code4biz_slack_messages.xlsx'
    wb = openpyxl.load_workbook(aggre_file_name, data_only=True)

    # 最後のシートを選択
    ws_summary = wb.worksheets[-1]

    # テーブルを生成する
    table = Table(displayName='T集計', ref='A1:I200')

    # テーブルのスタイルを決める(紫、テーブルスタイル(中間))
    table_style = TableStyleInfo(name='TableStyleMedium5', showRowStripes=True)

    # テーブルのスタイルを設定
    table.tableStyleInfo = table_style

    # シートにテーブルを追加
    ws_summary.add_table(table)

    # テーブル化した列幅の調整
    ws_summary.column_dimensions['A'].width = 16
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 25
    ws_summary.column_dimensions['G'].width = 20
    ws_summary.column_dimensions['H'].width = 50
    ws_summary.column_dimensions['I'].width = 50

    # テーブル化したフォントの設定
    font_name = 'Meiryo'

    # タイトル行のみ文字色を白色へ
    for i in range(1, ws_summary.max_row + 1):
        ws_summary.cell(row=1, column=i).font = Font(name=font_name, sz=11, color='FFFFFF')

    for i in range(2, ws_summary.max_row + 1):
        ws_summary.cell(row=i, column=1).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=2).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=3).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=4).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=5).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=6).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=7).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=8).font = Font(name=font_name, sz=11)
        ws_summary.cell(row=i, column=9).font = Font(name=font_name, sz=11)

    wb.save(aggre_file_name)


def main(start, end):
    # 抽出期間の取得
    search_days = create_search_day_list(start, end)

    # 抽出期間より1日毎取り出す
    for _search_day in search_days:
        search_day = _search_day.strftime("%Y-%m-%d")

        # データ格納用
        data = []

        driver = selenium_run()
        auto_login(driver)

        # 抽出日で全件検索
        to_search_date(search_day)
        sleep(2)

        # 検索件数の取得
        search_results_count = int(driver.find_element(by=By.CLASS_NAME, value='c-tabs__tab_count').text)
        print(f'検索結果数：{search_results_count}')

        # 並べ替え：新 → 古
        pg.press('tab', presses=16, interval=0.2)
        pg.press('enter', presses=1, interval=0.1)
        pg.press('down', presses=2, interval=0.2)
        pg.press('enter', presses=1, interval=0.1)
        # 全件表示
        pg.press('tab', presses=1, interval=0.2)
        pg.press('enter', presses=1, interval=0.1)
        pg.press('down', presses=3, interval=0.2)
        pg.press('enter', presses=1, interval=0.1)

        # 表示を小さくする
        display_smaller()

        # 一番下まで移動
        pg.press('tab', presses=3, interval=0.5)
        pg.press('pagedown', presses=10, interval=0.5)
        sleep(2)

        # 取得日時
        start_time = dt.now()

        # メッセージ重複有り（取得後、pandasにて重複削除）
        roop_total = math.floor(search_results_count / 21)
        for i in range(roop_total + 1):  # roop_total 回までスクロール
            message_groups = driver.find_elements(by=By.CLASS_NAME, value='c-message_group')
            print(f'メッセージ数:{len(message_groups)}')

            # 最後（検索結果の下側）から取得するためリスト反転
            message_groups.reverse()

            for j, message in enumerate(message_groups):
                date = message.find_element(by=By.CLASS_NAME, value='c-message_group__header_date').text
                timestamp = message.find_element(by=By.CLASS_NAME, value='c-timestamp__label').text
                channel_name = message.find_element(by=By.CLASS_NAME, value='p-deprecated_channel_name__text').text
                sender_name = message.find_element(by=By.CLASS_NAME, value='c-message__sender').text

                # 長文の場合の”...もっと表示する”となっている場合の対応
                try:
                    if len(message.find_elements(by=By.CLASS_NAME, value='c-search__expand_ellipsis')) > 0:
                        print('”...もっと表示する”あり')
                        message.find_element(by=By.CLASS_NAME, value='c-search__expand_ellipsis').click()
                        sleep(1)
                except:
                    pass

                try:
                    text_section = get_text_section(message, class_name='p-rich_text_section')
                except:
                    try:
                        text_section = get_text_section(message, class_name='c-search_message__body')
                    except:
                        print('除外処理')
                        text_section = ''  # エクセルで検索しやすいように空白

                urls = message.find_elements(by=By.TAG_NAME, value='a')
                url = urls[1].get_attribute('href')
                thread_ts = url.split('=')[-1]

                datum = {
                    '検索結果数': search_results_count,
                    '抽出日時': start_time,
                    'thread_ts': thread_ts,
                    '投稿日': date,
                    '投稿時間': timestamp,
                    '投稿チャンネル': channel_name,
                    '投稿者': sender_name,
                    '投稿メッセージ': text_section,
                    'リンク': url
                }
                data.append(datum)
                print(
                    f'{search_day} {i}/{roop_total}スクロール目{j + 1}/{len(message_groups)}件 (全{search_results_count}件): {sender_name} {timestamp}「{text_section[:10]} ・・・」メッセージ取得完了')
                # 待機時間（サイトに負荷を与えないと同時にコンテンツの読み込み待ち）
                sleep(1)

            # リスト反転もどす
            message_groups.reverse()

            # 検索結果より1度の取得件数が少なければ、取得メッセージ先頭の要素までスクロール
            if search_results_count > len(message_groups):
                first_message = message_groups[0]
                actions = ActionChains(driver)
                actions.move_to_element(first_message)
                actions.perform()
                # 待機時間（サイトに負荷を与えないと同時にコンテンツの読み込み待ち）
                sleep(3)
            else:
                break

            # 待機時間（サイトに負荷を与えないと同時にコンテンツの読み込み待ち）
            sleep(3)

        df = pd.DataFrame(data)
        # 重複行の削除
        df = df.drop_duplicates()
        df['リンク'] = df['リンク'].apply(make_clickable)

        # ソースディレクトリの作成
        src_dir = create_src_dir(search_day)

        # 検索件数と取得件数の不一致時に判別用ファイル名
        if df['検索結果数'].count() == search_results_count:
            aggre_file_name = f'{src_dir}/{search_day}_code4biz_slack_messages.xlsx'
        else:
            aggre_file_name = f'{src_dir}/【取得数不足】{search_day}_code4biz_slack_messages.xlsx'
        df.to_excel(aggre_file_name, index=False)

        excel_tabling(src_dir, search_day)
        print(f'{search_day}の取得完了')

        # 1日ごとにChrome終了
        driver.quit()


if __name__ == '__main__':
    # 期間指定
    start = '2022-03-15'
    end = '2022-03-31'

    main(start=start, end=end)
    print(f'{start}～{end}の全件取得完了')
